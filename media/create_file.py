from openpyxl import load_workbook, Workbook

def read_excel_and_save(input_file, output_file):
    try:
        # Cargar el archivo Excel existente
        wb = load_workbook(input_file)
        sheet = wb.active
        
        # Leer datos del archivo Excel y guardarlos en una lista
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        
        # Crear un nuevo workbook para guardar los datos
        wb_new = Workbook()
        sheet_new = wb_new.active
        
        # Escribir los datos en el nuevo workbook
        for row_idx, row_data in enumerate(data):
            for col_idx, col_value in enumerate(row_data):
                sheet_new.cell(row=row_idx + 1, column=col_idx + 1, value=col_value)
        
        # Guardar el nuevo archivo Excel
        wb_new.save(output_file)
        print(f"Nuevo archivo Excel guardado en: {output_file}")
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    # Ruta del archivo Excel existente
    input_file = 'excel.xlsx'
    
    # Ruta donde se guardará el nuevo archivo Excel
    output_file = 'output.xlsx'
    
    # Llamar a la función para leer y guardar
    read_excel_and_save(input_file, output_file)