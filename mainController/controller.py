from django.db import transaction
from django.shortcuts import get_object_or_404
from multiprocessing import Pool
import routeros_api
import uuid
import socket
import json
from pprint import pprint
import time
from .models import Devices, GroupDevices
from .unifiApi import Unifi
from .ruckusApi import Ruckus
from .brocadeApi import Brocade
from .mikrotikApi import Mikrotik
from .vszApi import connectVsz
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from bson import ObjectId
import os

def distributor(test):
    ip_address, user, password, vendor, collection, state, device_type = (test[0], test[1], test[2], test[3], test[4].rstrip(), test[5], test[6])
    if vendor == 'unifi':
        device = connect_device(Unifi, ip_address, user, password, collection, vendor, state, device_type)
    elif vendor == 'ruckus':
        device = connect_device(Ruckus, ip_address, user, password, collection, vendor, state, device_type)
    elif vendor == 'brocade':
        device = connect_device(Brocade, ip_address, user, password, collection, vendor, state, device_type)
    elif vendor == 'mikrotik':
        print('mikrotik')
        device = connect_device(Mikrotik, ip_address, user, password, collection,vendor, state, device_type)
    return device

def checkHost(ip_address, vendor):
    if vendor == 'mikrotik':
        port = 8728
    else:
        port = 22
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.settimeout(0.5)
        return not sock.connect_ex((ip_address, port))

def connect_device(DeviceClass, ip_address, user, password, collection, vendor, state="configured", device_type="access_point"):
    print('intentando conectar a '+ip_address)
    if checkHost(ip_address, vendor):
        print('responde')
        device = DeviceClass(ip_address, user, password)
        if device.status == 1:
            print('device status 1*******')
            hostname = device.getDeviceName()
            group = GroupDevices.objects.get(group_name=collection)
            data = device.getData()
            mac_address = data.get('mac address', '')
            clients  = device.get_clients()
            device_data = {
                'group': group,
                'deviceUser': user,
                'devicePassword': device.password,
                'ipAddress': ip_address,
                'deviceName': hostname,
                'model': data.get('model', ''),
                'macAddress': mac_address,
                'vendor': vendor,
                'version': data.get('version', ''),
                'controllerStatus': 'null',
                'clientes': clients,
                'state': state,
                'serialNumber':data.get('serial', ''),
                'status': device.status,
                'deviceType':device_type,
            }
            print(device_data)
            Devices.objects.update_or_create(ipAddress=ip_address.strip(), defaults=device_data)
            return device.status
        else:
            # print(ip_address+'no responde')
            # device_data ={
            #     'ipAddress':ip_address,
            #     'status':1,
            #     'deviceUser': user,
            #     'devicePassword': device.password,
            # }
            # Devices.objects.update_or_create(ipAddress=ip_address, defaults=device_data)
            return device.status
    else:
        return '2'


def connect_device_update(DeviceClass, ip_address, user, password, collection):
    if checkHost(ip_address):
        print(ip_address)
        device = DeviceClass(ip_address, user, password)
        if device.status == 1:
            interfaces = device.getInterfacesDevices()
            #print(data)
            device_data = {
                'clientes': interfaces,
            }
            print(device_data)
            Devices.objects.update_or_create(ipAddress=ip_address, defaults=device_data)

def scan_devices(devices):
    with transaction.atomic():
        for host in devices:
            distributor(host)


#######################################
#########INICIO FUNCIONES VSZ##########
#######################################

def set_ap_controller(devices):
    for host in devices:
        print(host)
        device = Ruckus(host[0], host[1], host[2])#### pasar id desde views.py
        if device.status == 1:
            hostname = device.setController()
            _id = host[4].strip()
            Devices.objects.filter(_id=ObjectId(_id)).update(state='oncontroller')
            device.sendCommand('exit')
    return "ok" 

def get_device_from_vsz(mac_address, id):
    vsz = connectVsz(mac_address=mac_address)
    ap_info = vsz.get_ap_info()
    json_ap_info = json.loads(ap_info)
    print(json_ap_info.get("name"))
    ip_address = json_ap_info.get("network", {}).get("ip")
    nodo_padre = {}
    # if checkHost(ip_address, "ruckus"):
    device = get_object_or_404(Devices, ipAddress='192.168.188.7')
    dict_clients = device.clientes
    for interface, mac in dict_clients.items():
        if mac.upper() == mac_address:
            nodo_padre = {'puerto':interface}

    Devices.objects.filter(_id=id).update(
        ipAddress = ip_address,
        deviceName = json_ap_info.get("name"),
        controllerStatus = json_ap_info.get("description"),
        clientes = [nodo_padre]
        )
    return "ok"

def update_info_from_excel(mac_address, serial):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    MEDIA_DIR = os.path.join(BASE_DIR, 'media')

    nombre_archivo = 'output.xlsx'
    file_path = os.path.join(MEDIA_DIR, nombre_archivo)

    if not os.path.exists(file_path):
        print(f"Error: The file '{file_path}' does not exist.")
        return ['error', 'error', 'error']

    try:
        workbook = load_workbook(file_path, keep_vba=True)
        worksheet = workbook.active

        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append(list(row))

        updated = False
        for index, row in enumerate(data):
            if len(row) >= 6:
                if row[5] != 'en uso':
                    print(row[5])
                    worksheet.cell(row=index + 1, column=3, value=mac_address)
                    worksheet.cell(row=index + 1, column=4, value=serial)
                    worksheet.cell(row=index + 1, column=6, value='en uso')
                    name, ip, mac = row[0], row[1], row[4]
                    updated = True
                    break
            else:
                print("Row does not have enough elements")

        if updated:
            workbook.save(file_path)
            workbook.close()
            return name, ip, mac
        else:
            print("No rows were updated.")
            workbook.close()
            return ['error', 'error', 'error']

    except InvalidFileException as e:
        print(f"Invalid file format: {e}")
        return ['error', 'error', 'error']
    except Exception as e:
        print(f"Error loading or saving workbook: {e}")
        return ['error', 'error', 'error']

def put_ap_info_on_vsz(devices):
    for device in devices:
        mac_address = device[0]
        hostname, ip_address, description = update_info_from_excel(device[0], device[1])
        print(hostname, ip_address, description)
        new_ap = connectVsz(mac_address)
        if new_ap.search_ap() == "ok":
            new_ap.config_ap(hostname=hostname,ip_address=ip_address,description=description)
    return 'ok'

####################################
#########FIN FUNCIONES VSZ##########
####################################

def update_device_info(ip_address, user, password, collection):
    connect_device_update(Brocade, ip_address, user, password, collection)
    return "ok"

def connect_mikrotik(device):
    print(device)
    connection = routeros_api.RouterOsApiPool(device[0], username=device[1], password=device[2])
    api = connection.get_api()
    list_dhcp = api.get_resource('/ip/dhcp-server/lease')
    routerboard = api.get_resource('/system/routerboard')
    identity = api.get_resource('/system/identity')    
    _identity = identity.get()    
    interfaces = api.get_resource('/interface/ethernet')
    _routerboard = routerboard.get()
    _identity = identity.get()
    for interface in interfaces.get():
        if interface['name']=='ether1':
            mac_address=interface['mac-address']
    version =_routerboard[0]['current-firmware']
    model =_routerboard[0]['model']
    dhcp_list = list_dhcp.get()
    group = GroupDevices.objects.get(group_name=device[3])
    list_client = []
    for dhcp_client in dhcp_list:
        if 'mac-address' in dhcp_client and dhcp_client['mac-address']:
            list_client.append({
                'address':dhcp_client['address'],
                'mac':dhcp_client['mac-address'],
                'server':dhcp_client['server']
                })
    device_data = {
        'group': group,
        'deviceUser': device[1],
        'devicePassword': device[2],
        'ipAddress': device[0],
        'deviceName': _identity[0]['name'],
        'model': model,
        'macAddress': mac_address,
        'version': version,
        'controllerStatus': 'null',
        'clientes': list_client,
        'status': 2,
    }
    Devices.objects.update_or_create(ipAddress=device[0], defaults=device_data)


def main():
    start = time.time()
    hosts = getHosts()
    pprint(hosts)
    with transaction.atomic():
        for host in hosts:
             distributor(host)
    finish = time.time()
    print(float("{:.2f}".format(finish - start)))

if __name__ == "__main__":
    main()
