from django.conf import settings
from openpyxl import load_workbook
from unificontrol import UnifiClient
import os
from pprint import pprint
import routeros_api
import socket

def _read_excel():
    base_dir = settings.BASE_DIR
    print(settings.BASE_DIR)
    file_path = os.path.join(base_dir, 'mainController/excels', 'test_excel.xlsx')

    workbook = load_workbook(file_path)
    worksheet = workbook.active
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(list(row))

    print(data[0][1])

    workbook.close()
    return(data)

def unifi_controller(ip_controller, site):
    uc_user = "rsupport"
    uc_pass = "elrbsestNF!25"
    uc_site = site
    client = UnifiClient(host=ip_controller,
    username=uc_user, password=uc_pass, site=uc_site)
    devices = client.list_devices()
    list_devices = []
    for device in devices:
        if device['ip'] == "192.168.200.100":
            list_devices.append({"ip":device['ip'],"id":device['_id'],"mac":device['mac']})
            pprint(device)
    pprint(list_devices)
    return(list_devices)

def connect_mikrotik():
    connection = routeros_api.RouterOsApiPool('50.50.50.1', username='admin', password='elrbsest')
    api = connection.get_api()
    list_dhcp = api.get_resource('/ip/dhcp-server/lease')
    dhcp_list = list_dhcp.get(server="dhcp1")
    list_client = []
    for dhcp_client in dhcp_list:
        # list_client.append(dhcp_client['address'])
        # list_client.append(dhcp_client['mac-address'])
        # list_client.append(dhcp_client['server'])
        list_client.append([dhcp_client['address'],dhcp_client['mac-address'],dhcp_client['server']])
    print(list_client[0][0])
    return(list_client)

def checkhost(vendor='mikrotik',ip_address='172.19.60.3' ):
    if vendor == 'mikrotik':
        port = 8728
    else:
        port = 22
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.settimeout(0.5)
        return not sock.connect_ex((ip_address, port))

if __name__ == "__main__":
    # _read_excel()
    if checkhost():
        print('responde')
    else:
        print('no responde')
